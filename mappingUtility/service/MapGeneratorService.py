from mappingUtility.Utility import ComponentUtils
from mappingUtility.client import BoomiApiService
from mappingUtility.service import MappingComponentXmlGenerator
from mappingUtility.strategy.JsonComponentGenerator import JSONProcessor
from mappingUtility.strategy.XmlComponentGenerator import XMLProcessor
from mappingUtility.strategy.EdifactComponentGenerator import EDIFACTProcessor
from mappingUtility.strategy.ComponentGeneratorContext import FileProcessingContext
import logging
import openpyxl
import io

logger = logging.getLogger(__name__)

def profile_component_generator(file_content,file_type):
        if file_type == 'edi' or file_type == 'EDIFACT' or file_type == 'edifact':
            strategy = EDIFACTProcessor()
        elif file_type == 'json' or file_type == 'JSON':
            strategy = JSONProcessor()
        elif file_type == 'xml' or file_type == 'XML':
            strategy = XMLProcessor()
        else:
            raise ValueError("Unsupported file type. Please provide 'json' or 'xml'.")

        context = FileProcessingContext(strategy)

        xml_response = context.execute(file_content)
        return xml_response

def read_excel_with_filetypes(excel_data_bytes):
    excel_file_like = io.BytesIO(excel_data_bytes)

    workbook = openpyxl.load_workbook(excel_file_like, data_only=True)

    source_info = None
    target_info = None

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if sheet.sheet_state == 'hidden':
            sheet.sheet_state = 'visible'

        if sheet_name.startswith("source_"):
            filetype = sheet_name.split("source_")[-1]
            value = sheet["A2"].value
            source_info = {"sheet": sheet_name, "filetype": filetype, "value": value}
        elif sheet_name.startswith("target_"):
            filetype = sheet_name.split("target_")[-1]
            value = sheet["A2"].value
            target_info = {"sheet": sheet_name, "filetype": filetype, "value": value}

    if not source_info or not target_info:
        missing_sheets = [sheet_name for sheet_name in workbook.sheetnames if not sheet_name.startswith("source_") and not sheet_name.startswith("target_")]
        logger.error(f"Missing required sheets. Available sheets: {workbook.sheetnames}. Missing sheets: {missing_sheets}")
        raise ValueError("Missing sheet with 'source_' or 'target_' prefix")

    return source_info, target_info

def generate_component_xml_with_excel( excel_data):
    try:
        source_info, target_info = read_excel_with_filetypes(excel_data)
        
        source_data = source_info['value']
        source_file_type = source_info['filetype']
        destination_data = target_info['value']
        destination_file_type = target_info['filetype']
        
        logger.info("Starting process_mapping_files execution.")
        
        sourceTempXml = profile_component_generator(source_data, source_file_type)
        sourceXml = BoomiApiService.upload_xml_component_to_boomi(sourceTempXml)
        sourceComponentId = ComponentUtils.extract_component_id(sourceXml)
        logger.info(f"Extracted source component ID: {sourceComponentId}")
        
        destinationTempXml = profile_component_generator(destination_data, destination_file_type)
        destinationXml = BoomiApiService.upload_xml_component_to_boomi(destinationTempXml)
        destinationComponentId = ComponentUtils.extract_component_id(destinationXml)
        logger.info(f"Extracted destination component ID: {destinationComponentId}")
        
        mapTempComponentXml = MappingComponentXmlGenerator.generate_boomi_map(
            excel_data,
            sourceXml,
            destinationXml,
            source_col="Source Field (Dropdown)",
            target_col="Target Field",
            from_profile_id=sourceComponentId,
            to_profile_id=destinationComponentId
        )
        logger.info("Generated Boomi map component XML.")
        
        mapComponentXml = BoomiApiService.upload_xml_component_to_boomi(mapTempComponentXml)
        logger.info("Uploaded map component.")
        
        return mapComponentXml

    except Exception as e:
        logger.error(f"An error occurred during process_mapping_files execution: {e}")
        raise
