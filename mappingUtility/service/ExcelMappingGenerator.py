import json
import xml.etree.ElementTree as ET
import os
import io
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font
import re
from collections import Counter
import posixpath
import requests
import traceback
from django.http import HttpResponse
from resources.globlas import GEN_AI_API_KEY,GEN_AI_URL
import logging
from mappingUtility.Utility import EdifactUtils

os.path = posixpath

logger = logging.getLogger(__name__)

def read_file_from_resources(file_path):
    with open(file_path, 'r') as file:
        return file.read()
    
def read_edi_promptes():
    resource_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'resources', 'ediPrompt.txt')
    template = read_file_from_resources(resource_path)
    return template

def load_main_prompts(source_fields, target_fields):
    resource_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'resources', 'mainPrompt.txt')
    template = read_file_from_resources(resource_path)

    source_str = ', '.join(source_fields)
    target_str = ', '.join(target_fields)

    formatted_template = template.format(sourceNames=source_str, targetNames=target_str)
    return formatted_template

def extract_json_paths(data, parent_key="", seen_paths=None):
    if seen_paths is None:
        seen_paths = set()
    paths = []
    
    if isinstance(data, dict):
        for key, value in data.items():
            new_key = f"{parent_key}.{key}" if parent_key else key
            if isinstance(value, (dict, list)):
                paths.extend(extract_json_paths(value, new_key, seen_paths))
            else:
                if new_key not in seen_paths:
                    seen_paths.add(new_key)
                    paths.append(new_key)
    elif isinstance(data, list):
        new_key = f"{parent_key}[*]" if parent_key else "[*]"
        if new_key not in seen_paths:
            seen_paths.add(new_key)
            if data:
                paths.extend(extract_json_paths(data[0], new_key, seen_paths))
    
    return paths

def extract_xml_paths(element, parent_path="", seen_paths=None):
    if seen_paths is None:
        seen_paths = set()
    paths = []
    
    current_path = f"{parent_path}/{element.tag}" if parent_path else element.tag
    if list(element):
        for child in element:
            paths.extend(extract_xml_paths(child, current_path, seen_paths))
    else:
        if current_path not in seen_paths:
            seen_paths.add(current_path)
            paths.append(current_path)
    
    return paths

def extract_edifact_fields(edifact_content, seg_sep="'", elem_sep="+", sub_elem_sep=":"):
    resource_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'resources', 'D96_ORDERS.xml')
    xml_content = read_file_from_resources(resource_path)
    return EdifactUtils.get_edifact_fields(xml_content, edifact_content)

def extract_x12_fields(content, seg_sep="~", elem_sep="*", sub_elem_sep=":"):
    fields = []
    seen_paths = set()
    segment_counts = Counter()

    segments = content.strip().split(seg_sep)
    
    for segment in segments:
        parts = segment.strip().split(elem_sep)
        segment_type = parts[0].strip()
        segment_counts[segment_type] += 1

    for segment in segments:
        segment = segment.strip()
        if not segment:
            continue

        parts = segment.split(elem_sep)  
        segment_type = parts[0].strip() 
        qualifier = parts[1].strip() if len(parts) > 1 else ""

        for i, element in enumerate(parts[1:], start=1):
            sub_elements = element.split(sub_elem_sep) 
            field_index = str(i)
            if i < 10:
                field_index = "0" + str(i)
                
            if len(sub_elements) > 1:
                for j, sub_element in enumerate(sub_elements, start=1):
                    field_path = f"{segment_type}{field_index}.{j}_{qualifier}" if segment_counts[segment_type] > 1 else f"{segment_type}{field_index}.{j}"
                    if field_path not in seen_paths and sub_element.strip():
                        seen_paths.add(field_path)
                        fields.append(field_path)
            else:
                field_path = f"{segment_type}{field_index}_{qualifier}" if segment_counts[segment_type] > 1 else f"{segment_type}{field_index}"
                if field_path not in seen_paths:
                    seen_paths.add(field_path)
                    fields.append(field_path)
    return fields

def read_content(content, file_type, seg_sep="~", elem_sep="*", sub_elem_sep=":"):
    if file_type == "JSON":
        try:
            json_content = json.loads(content);
            return extract_json_paths(json_content)
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON: {e}")
    elif file_type == "XML":
        try:
            return extract_xml_paths(ET.fromstring(content))
        except ET.ParseError as e:
            raise ValueError(f"Invalid XML: {e}")
    elif file_type == "EDIFACT":
        return extract_edifact_fields(content, seg_sep="'", elem_sep="+", sub_elem_sep=":")
    elif file_type == "X12":
        return extract_x12_fields(content, seg_sep, elem_sep, sub_elem_sep)
    else:
        raise ValueError("Unsupported file type")

def call_gemini_api(prompt):
    try:
        response = requests.post(
            GEN_AI_URL + GEN_AI_API_KEY,
            headers={"Content-Type": "application/json"},
            json={
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": {"temperature": 0.3, "maxOutputTokens": 65536}
            }
        )
        response_json = response.json()
        ai_response = response_json["candidates"][0]["content"]["parts"][0]["text"]
        cleaned_json = re.sub(r"```json\s*|\s*```", "", ai_response).strip()
        return json.loads(cleaned_json)
    except Exception as e:
        logger.error(f"API Error: {str(e)}")
        return {}



def generate_mapping_prompt(source_fields, target_fields):
    prompt_template = load_main_prompts(source_fields, target_fields)
    return prompt_template

def get_mapping_from_ai(source_fields, target_fields):
    try:
        prompt = generate_mapping_prompt(source_fields, target_fields);
        return call_gemini_api(prompt)
    except Exception as e:
        logger.error(f"Mapping Error: {str(e)}")
        return {}

def normalize_confidence(confidence):
    """Ensure confidence is always in percentage format (0-100)."""
    if confidence is None:
        return 0.0
    elif confidence > 1:
        return round(confidence, 2)
    else:
        return round(confidence * 100, 2)

def generate_edi_description_prompt(edi_fields):
    prompt = read_edi_promptes();
    ediFields = json.dumps(edi_fields, indent=2)
    prompt = prompt.format(ediFields=ediFields)
    return prompt

def get_edi_field_descriptions(edi_fields):
    description_prompt = generate_edi_description_prompt(edi_fields)
    return call_gemini_api(description_prompt)

def create_excel_mapping(field_mappings, source_fields, target_fields, source_format, target_format, excel_filename="AI_Field_Mapping.xlsx"):
    """
    Creates an Excel file with field mappings, confidence levels, and dropdowns for source fields.

    Parameters:
        field_mappings (dict): A dictionary containing mappings between target and source fields.
        source_fields (list): A list of source fields to populate the dropdown.
        target_fields (list): A list of target fields to map.
        source_format (str): The format of the source data (e.g., "json", "xml", "edifact", "x12").
        target_format (str): The format of the target data (e.g., "json", "xml", "edifact", "x12").
        excel_filename (str): The name of the Excel file to save. Defaults to "AI_Field_Mapping.xlsx".
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Field Mapping"
    headers = ["Target Field", "Source Field (Dropdown)", "Confidence (%)", "Mapping Type", "Logic"]
    ws.append(headers)

    # Apply bold formatting to headers
    bold_font = Font(bold=True)
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header).font = bold_font

    # Add dropdown options for Mapping Type
    mapping_type_dv = DataValidation(type="list", formula1='"Direct,Logic"', showDropDown=True)

    # Create a hidden sheet for dropdown values (fixes 255-character limit issue)
    ws_hidden = wb.create_sheet(title="DropdownValues")
    extra_options = ["<Not Mapped>", "<Constant>", "<No Mapping Needed>", "<Logic>"]

    # Define color fills for confidence levels
    high_conf_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Green
    medium_conf_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    low_conf_fill = PatternFill(start_color="FF5050", end_color="FF5050", fill_type="solid")  # Red

    # Write Source Fields to the hidden sheet
    ws_hidden["A1"] = "Source Fields"
    actual_source_fields = [str(field) for field in source_fields]
    actual_target_fields = [str(field) for field in target_fields]
    total_dropdown_fields = extra_options + actual_source_fields

    for i, path in enumerate(total_dropdown_fields, start=2):
        ws_hidden[f"A{i}"] = path

    # Define the range for dropdown values
    dropdown_range = f"'DropdownValues'!$A$2:$A${len(total_dropdown_fields) + 1}"

    # Ensure field_mappings is a dictionary
    if not isinstance(field_mappings, dict):
        raise TypeError("‚ùå Error: field_mappings should be a dictionary!")

    # Process field mappings and generate Excel
    for row_num, target_field in enumerate(target_fields, start=2):
        mapping = field_mappings.get(target_field, None)
        if not isinstance(mapping, dict):
            mapping = {}

        if mapping:
            source_field = mapping.get("source_field", "<Not Mapped>")
            confidence = float(mapping.get("confidence", 0.0))
            confidence = normalize_confidence(mapping.get("confidence", 0))
            mapping_type = mapping.get("mapping_type", "Direct")
            logic = mapping.get("logic", "")
        else:
            source_field = "<Not Mapped>"
            confidence = 0
            mapping_type = "Manual"
            logic = "Needs manual review"

        # Create dropdown validation
        dv = DataValidation(type="list", formula1=dropdown_range, showDropDown=True)
        ws.add_data_validation(dv)

        # Write AI-mapped source field with dropdown support
        source_cell = ws.cell(row=row_num, column=2)
        if isinstance(source_field, list):
            source_cell.value = ", ".join(source_field)
        else:
            source_cell.value = source_field if source_field else "Not Mapped"
        dv.add(source_cell)

        # Apply color coding based on confidence level
        if confidence >= 80:
            source_cell.fill = high_conf_fill
        elif confidence >= 50:
            source_cell.fill = medium_conf_fill
        else:
            source_cell.fill = low_conf_fill

        # Write target field and confidence score
        ws.cell(row=row_num, column=1, value=target_field)
        ws.cell(row=row_num, column=3, value=confidence)

        # Add "Mapping Type" dropdown (Direct / Logic)
        mapping_type_cell = ws.cell(row=row_num, column=4, value=mapping_type)
        mapping_type_dv.add(mapping_type_cell)
        ws.add_data_validation(mapping_type_dv)

        # Write business logic if applicable
        ws.cell(row=row_num, column=5, value=logic)

    # Hide the dropdown values sheet
    ws_hidden.sheet_state = "hidden"

    if source_format in ["EDIFACT", "X12"] or target_format in ["EDIFACT", "X12"]:
        if source_format in ["edifact", "x12"]:
            edi_fields = source_fields
        else:
            edi_fields = target_fields

        # Create a new sheet for EDI Field Descriptions
        ws_desc = wb.create_sheet(title="EDI Field Descriptions")
        ws_desc.append(["Field Name", "Description"])

        # Fetch AI-generated descriptions
        edi_descriptions = get_edi_field_descriptions(edi_fields)
        if edi_descriptions:
            for field, description in edi_descriptions.items():
                ws_desc.append([field, description])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)  # Important: reset buffer position
    return output

    

def main(source_type, source_data, target_type, target_data):
    logger.info(f"Source Type: {source_type}")
    logger.info(f"Target Type: {target_type}")
    # logger.info(f"Source Data: {source_data}")
    # logger.info(f"Target Data: {target_data}")
    try:
        source_fields = read_content(source_data, source_type)
        logger.info(f"Source Fields: {source_fields}")
        target_fields = read_content(target_data, target_type)
        logger.info(f"Target Fields: {target_fields}")
        mappings =  get_mapping_from_ai(source_fields, target_fields)
        logger.info(f"Mappings: {mappings}")
        output = create_excel_mapping(mappings, source_fields, target_fields, source_type, target_type)
        logger.info(f"Excel file created successfully")
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="AI_Field_Mapping.xlsx"'
        return response
        
    except Exception as e:
        tb = traceback.format_exc()
        error_response = json.dumps({
            "status": "error",
            "message": str(e),
            "trace": tb,
            "line": tb.splitlines()[-1] if tb else "Unknown"
        })
        return HttpResponse(error_response, content_type="application/json", status=500)

